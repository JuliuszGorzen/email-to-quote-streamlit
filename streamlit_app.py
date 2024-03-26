import datetime
import io
import json
import math
import os.path
import re
from collections import Counter

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
import streamlit_authenticator as stauth
import yaml
from annotated_text import annotated_text, annotation
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
from langchain.chains import create_retrieval_chain
from langchain.chains.combine_documents import create_stuff_documents_chain
from langchain_community.callbacks import get_openai_callback
from langchain_community.vectorstores import FAISS
from langchain_core.messages import SystemMessage, HumanMessage, AIMessage
from langchain_core.prompts import ChatPromptTemplate, SystemMessagePromptTemplate, HumanMessagePromptTemplate
from langchain_openai import AzureChatOpenAI, AzureOpenAIEmbeddings
from langchain_text_splitters import MarkdownHeaderTextSplitter
from pycm import ConfusionMatrix
from yaml.loader import SafeLoader

import constants


def main() -> None:
    create_page_config()

    credential = DefaultAzureCredential()
    secret_client = SecretClient(vault_url="https://genai-dev-keyvault.vault.azure.net/", credential=credential)
    openai_api_key_secret = secret_client.get_secret("aikey").value
    llm = create_azure_openai_model(openai_api_key_secret, 0.2)
    embedding_llm = create_azure_openai_embedding_model(openai_api_key_secret)

    hide_sidebar()
    create_version_label()

    authenticator = create_authenticator()
    create_login_page(authenticator)

    if st.session_state["authentication_status"]:
        display_sidebar()
        create_logout_button(authenticator)
        create_main_page()
        create_sidebar()
        create_tabs(llm, embedding_llm)
    elif st.session_state["authentication_status"] is False:
        st.error(constants.LOGIN_PAGE_ERROR_MESSAGE)
    elif st.session_state["authentication_status"] is None:
        st.warning(constants.LOGIN_PAGE_WARNING_MESSAGE)


def create_logout_button(authenticator: stauth.Authenticate) -> None:
    authenticator.logout(
        button_name=constants.LOGOUT_BUTTON_TEXT,
        location="sidebar"
    )


def create_login_page(authenticator: stauth.Authenticate) -> None:
    authenticator.login(fields={
        "Form name": constants.LOGIN_PAGE_NAME,
        "Username": "username",
        "Password": "password",
        "Login": constants.LOGIN_BUTTON_TEXT
    })


def create_main_page() -> None:
    st.title(constants.MAIN_PAGE_HEADER)

    with st.expander(constants.HOW_TO_START_EXPANDER):
        st.markdown(read_md_file("markdowns/how-to-start-description.md"))

    with st.expander(constants.IMPORTANT_FILES_EXPANDER):
        st.markdown(read_md_file("markdowns/important-files-description.md"))

        col1, col2, col3 = st.columns(3)

        with col1:
            with open("important-files/example-markdown-rag.md", "r", encoding="utf-8") as file:
                st.download_button(
                    label="Example .md file (RAG) :bookmark_tabs:",
                    data=file,
                    file_name="example-markdown-rag.md",
                    mime="text/markdown"
                )
            st.caption("Download the example .md file to use it in the RAG tab")

        with col2:
            with open("important-files/email-2-quote-dataset.xlsx", "rb") as file:
                st.download_button(
                    label="Dataset .xlsx file :bar_chart:",
                    data=file,
                    file_name="email-2-quote-dataset.xlsx",
                    mime="application/vnd.ms-excel"
                )
            st.caption("Download dataset file with sample emails")

            with open("important-files/confusion-matrix.xlsx", "rb") as file:
                st.download_button(
                    label="Confusion Matrix .xlsx file :bar_chart:",
                    data=file,
                    file_name="confusion-matrix.xlsx",
                    mime="application/vnd.ms-excel"
                )
            st.caption("Download confusion matrix file")

        with col3:
            with open("important-files/email-2-quote-openapi.json", "r", encoding="utf-8") as file:
                st.download_button(
                    label="OpenAPI .json schema :book:",
                    data=file,
                    file_name="email-2-quote-openapi.json",
                    mime="application/json"
                )
            st.caption("Download OpenAPI schema (JSON)")

            with open("important-files/email-2-quote-openapi.yaml", "r", encoding="utf-8") as file:
                st.download_button(
                    label="OpenAPI .yaml schema :book:",
                    data=file,
                    file_name="email-2-quote-openapi.yaml",
                    mime="application/x-yaml"
                )
            st.caption("Download OpenAPI schema (YAML)")

            with open("important-files/email-2-quote-openapi-one-line.json", "r", encoding="utf-8") as file:
                st.download_button(
                    label="OpenAPI .json schema :book:",
                    data=file,
                    file_name="email-2-quote-openapi-one-line.json",
                    mime="application/json"
                )
            st.caption("Download OpenAPI schema (JSON - one line)")

    with st.expander(constants.DATASETS_EXPANDER):
        df_emails = pd.read_excel("important-files/email-2-quote-dataset.xlsx", "mails")
        st.dataframe(df_emails, use_container_width=True)

        df_objects = pd.read_excel("important-files/email-2-quote-dataset.xlsx", "objects").astype(str)
        st.dataframe(df_objects, use_container_width=True)


def create_version_label():
    annotated_text(annotation("version", "0.0.7", "#FF4B4B", font_size="1.5rem", float="right", color="#FFFFFF"))


def create_sidebar() -> None:
    st.sidebar.divider()
    st.sidebar.header(constants.SIDEBAR_HEADER)
    st.sidebar.error(constants.SIDEBAR_ERROR_MESSAGE)
    st.sidebar.subheader(constants.SIDEBAR_SUBHEADER)

    with st.sidebar.form("model_parameters_form"):
        st.write(constants.SIDEBAR_FORM_DESCRIPTION)
        st.text_input(
            label=constants.SIDEBAR_FORM_AZURE_ENDPOINT,
            value="https://open-ai-resource-gen-ai.openai.azure.com/",
            type="password",
            disabled=True
        )
        st.text_input(
            label=constants.SIDEBAR_FORM_OPENAI_API_VERSION,
            value="2023-07-01-preview",
            disabled=True
        )
        openai_api_key_val = st.text_input(
            label=constants.SIDEBAR_FORM_OPENAI_API_KEY,
            value="<api-key>",
            type="password",
            help=constants.SIDEBAR_FORM_OPENAI_API_KEY_HELP,
            disabled=True
        )
        st.text_input(
            label=constants.SIDEBAR_FORM_OPENAI_API_TYPE,
            value="azure",
            disabled=True
        )
        st.text_input(
            label=constants.SIDEBAR_FORM_DEPLOYMENT_NAME,
            value="gpt-35-dev",
            disabled=True
        )
        st.text_input(
            label=constants.SIDEBAR_FORM_MODEL_NAME,
            value="gpt-35-turbo",
            disabled=True
        )
        st.text_input(
            label=constants.SIDEBAR_FORM_MODEL_VERSION,
            value="0613",
            disabled=True
        )
        temperature_slider_val = st.slider(
            label=constants.SIDEBAR_FORM_TEMPERATURE,
            min_value=0.0,
            max_value=1.0,
            value=0.2,
            step=0.01,
            help=constants.SIDEBAR_FORM_TEMPERATURE_HELP,
            disabled=True
        )
        submitted = st.form_submit_button(
            label=constants.SIDEBAR_FORM_SUBMIT_BUTTON,
            on_click=st.balloons,
            use_container_width=True,
            disabled=True
        )

        if submitted:
            st.write(constants.SIDEBAR_FORM_SUMMARY)
            st.write("Temperature:", temperature_slider_val)
            st.write("OpenAI API key:", openai_api_key_val)

    with st.sidebar.expander(constants.SIDEBAR_FORM_MODEL_DESCRIPTION):
        st.markdown(read_md_file("markdowns/model-description.md"))


def create_tabs(llm: AzureChatOpenAI, embedding_llm: AzureOpenAIEmbeddings) -> None:
    zero_shot_prompting_tab, few_shot_prompting_tab, ner_zero_shot_prompting, ner_few_shot_prompting, rag, test_prompt = st.tabs(
        [
            constants.TAB_NAME_ZERO_SHOT_PROMPTING,
            constants.TAB_NAME_FEW_SHOT_PROMPTING,
            constants.TAB_NAME_NER_ZERO_SHOT_PROMPTING,
            constants.TAB_NAME_NER_FEW_SHOT_PROMPTING,
            constants.TAB_NAME_RAG,
            constants.TAB_NAME_TEST_PROMPT
        ])

    create_zero_shot_prompting_tab(zero_shot_prompting_tab, llm)
    create_few_shot_prompting_tab(few_shot_prompting_tab, llm)
    create_ner_zero_shot_prompting_tab(ner_zero_shot_prompting, llm)
    create_ner_few_shot_prompting_tab(ner_few_shot_prompting, llm)
    create_rag_tab(rag, llm, embedding_llm)
    create_test_prompt_tab(test_prompt, llm)


def create_test_prompt_tab(test_prompt: str, llm: AzureChatOpenAI) -> None:
    with test_prompt:
        st.header(constants.TEST_PROMPT_TAB_HEADER)

        with st.expander(constants.TAB_DESCRIPTION_EXPANDER_TEXT, expanded=True):
            st.markdown(read_md_file("markdowns/test-your-prompt-description.md"))

        with st.expander(constants.TAB_EXAMPLE_EXPANDER_TEXT):
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("Actual")
                df1 = pd.read_excel("important-files/confusion-matrix.xlsx", "actual").astype(str)
                st.dataframe(df1, use_container_width=True, height=200)

            with col2:
                st.markdown("Predicted")
                df2 = pd.read_excel("important-files/confusion-matrix.xlsx", "few-shot-prompting-predicted").astype(str)
                st.dataframe(df2, use_container_width=True, height=200)

            st.markdown("Confusion matrix")
            df3 = pd.read_excel("important-files/confusion-matrix.xlsx", "few-shot-prompting-cm")
            st.dataframe(df3, use_container_width=True, height=200)

        with st.form("confusion_matrix_form"):
            st.header(constants.CONFUSION_MATRIX_FORM_HEADER)

            col1, col2, col3 = st.columns(3)

            with col1:
                excel_file = st.file_uploader(
                    label=constants.TAB_FORM_EXCEL_FILE,
                    type=["xlsx"]
                )

            with col2:
                actual_sheet_name = st.text_input(
                    label=constants.TAB_FORM_ACTUAL_SHEET_NAME,
                    value="actual"
                )

            with col3:
                predicted_sheet_name = st.text_input(
                    label=constants.TAB_FORM_PREDICTED_SHEET_NAME,
                    value="predicted"
                )

            submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON_CONFUSION_MATRIX)

            if submitted:
                if excel_file is None:
                    st.warning("Upload an Excel file!")
                    st.stop()

                if not actual_sheet_name or not predicted_sheet_name:
                    st.warning("Fill in the sheet names!")
                    st.stop()

                with st.spinner("Calculating... :abacus:"):
                    saved_excel_path = save_excel_file(excel_file)
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    create_sheet_for_cm(saved_excel_path, actual_sheet_name, predicted_sheet_name, timestamp)
                    create_cm(saved_excel_path, actual_sheet_name, timestamp)

        # with st.form("test_your_prompt_form"):
        #     st.header(constants.TEST_PROMPT_TAB_FORM_HEADER)
        #      st.warning(":building_construction: This form is under construction. It doesn't work properly yet. "
        #                 ":building_construction:")
        #
        #     df_emails = pd.read_excel("important-files/email-2-quote-dataset.xlsx", "mails")
        #     st.dataframe(df_emails, use_container_width=True, height=200)
        #
        #     prompt_text = st.text_area(
        #         label=constants.TAB_FORM_PROMPT_MESSAGE,
        #         placeholder=constants.TEST_PROMPT_TAB_PROMPT,
        #         height=200,
        #         disabled=True
        #     )
        #
        #     submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON, disabled=True)
        #
        #     if submitted:
        #         if prompt_text == "":
        #             st.warning("Fill in the prompt field!")
        #             st.stop()
        #
        #         with st.status("Calculating... :abacus:", expanded=True):
        #             timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        #
        #             st.write("Running prompt against dataset... :running:")
        # prompt_parts = re.split("System:|Human:|AI:", prompt_text)
        # filter(None, prompt_parts)
        # emails_to_predict = df_emails["llm_input"].to_list()
        # llm_responses = []
        #
        # for email in emails_to_predict:
        #     prompt = ChatPromptTemplate.from_messages([
        #         SystemMessage(prompt_parts[0]),
        #         HumanMessage(prompt_parts[1]),
        #         AIMessage(prompt_parts[2]),
        #         HumanMessage(prompt_parts[3]),
        #         AIMessage(prompt_parts[4]),
        #         HumanMessage(prompt_parts[5]),
        #         AIMessage(prompt_parts[6]),
        #         HumanMessage(email)
        #     ])
        #     chain = prompt | llm
        #     response = chain.invoke({})
        #
        #     llm_responses.append(response.content)
        #
        #     for response in llm_responses:
        #         json_response: str
        #         try:
        #             response = json.loads(response)
        #         except json.JSONDecodeError:
        #             response = ""
        #
        #         if not any(col in response for col in df_emails.columns):
        #             llm_responses.remove(response)
        #
        # st.write("Preparing data... :eyes:")
        # create_sheet_for_cm("actual", "few-shot-prompting-predicted", timestamp)
        #
        # st.write("Preparing confusion matrix... :bar_chart:")
        # create_cm(timestamp)


def create_rag_tab(rag: str, llm: AzureChatOpenAI, embedding_llm: AzureOpenAIEmbeddings) -> None:
    with rag:
        st.header(constants.RAG_TAB_HEADER)

        with st.expander(constants.TAB_DESCRIPTION_EXPANDER_TEXT, expanded=True):
            st.markdown(read_md_file("markdowns/rag-description.md"))

        with st.expander(constants.TAB_EXAMPLE_EXPANDER_TEXT):
            st.markdown(read_md_file("markdowns/rag-example.md"))

        with st.expander(constants.TAB_FORM_EXPANDER_TEXT):
            with st.form("rag_form", border=False):
                st.header(constants.RAG_TAB_FORM_HEADER)
                system_message = st.text_area(
                    label=constants.TAB_FORM_SYSTEM_MESSAGE,
                    value=constants.RAG_TAB_SYSTEM_MESSAGE,
                    placeholder=constants.RAG_TAB_SYSTEM_MESSAGE,
                    height=200
                )

                col1, col2 = st.columns(2)

                with col1:
                    human_message = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        value=constants.RAG_TAB_HUMAN_MESSAGE,
                        placeholder=constants.RAG_TAB_HUMAN_MESSAGE,
                        height=200
                    )

                with col2:
                    external_file = st.file_uploader(
                        label=constants.TAB_FORM_FILE,
                        type=["md"]
                    )

                submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON)

                if submitted:
                    if is_any_field_empty([system_message, human_message]) or external_file is None:
                        st.warning(constants.TAB_FORM_EMPTY_FIELD_WARNING)
                        st.stop()

                    with get_openai_callback() as callbacks:
                        with st.spinner("Generating response..."):
                            headers_to_split_on = [
                                ("#", "Header 1"),
                                ("##", "Header 2"),
                                ("###", "Header 3"),
                                ("####", "Header 4")
                            ]
                            document = external_file.read().decode("utf-8")
                            markdown_splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
                            docs = markdown_splitter.split_text(document)
                            vectorstore = FAISS.from_documents(docs, embedding_llm)
                            prompt = ChatPromptTemplate.from_messages([
                                SystemMessagePromptTemplate.from_template(system_message),
                                HumanMessagePromptTemplate.from_template("{input}")
                            ])
                            document_chain = create_stuff_documents_chain(llm, prompt)
                            retriever = vectorstore.as_retriever()
                            retrieval_chain = create_retrieval_chain(retriever, document_chain)
                            response = retrieval_chain.invoke({"context": docs, "input": human_message})

                        st.markdown(constants.TAB_FORM_BOT_RESPONSE)

                        try:
                            st.json(json.loads(response["answer"]))
                        except ValueError:
                            st.text(response["answer"])

                        st.markdown(constants.TAB_FORM_FULL_PROMPT)
                        st.code(
                            prompt.format(context=[d.page_content for d in docs][0], input=human_message),
                            language="text"
                        )
                        st.markdown(constants.TAB_FORM_REQUEST_STATS)
                        st.text(callbacks)
                        create_success_toast()


def create_ner_few_shot_prompting_tab(ner_few_shot_prompting: str, llm: AzureChatOpenAI) -> None:
    with ner_few_shot_prompting:
        st.header(constants.NER_FEW_SHOT_PROMPTING_TAB_HEADER)

        with st.expander(constants.TAB_DESCRIPTION_EXPANDER_TEXT, expanded=True):
            st.markdown(read_md_file("markdowns/ner-few-shot-prompting-description.md"))

        with st.expander(constants.TAB_EXAMPLE_EXPANDER_TEXT):
            st.markdown(read_md_file("markdowns/ner-few-shot-prompting-example.md"))
            annotated_text(
                "Hello, Please send me your offer for groupage transport for: 1 pallet: ",
                ("120cm x 80cm x 120cm", "load_type"),
                ("155 Kg", "weight"),
                "Loading: ",
                ("300283 Timisoara, Romania", "origin_location"),
                "Unloading: ",
                ("4715-405 Braga, Portugal", "destination_location"),
                "Can be picked up. Payment after 7 days"
            )

        with st.expander(constants.TAB_FORM_EXPANDER_TEXT):
            with st.form("ner_few_shot_prompting_form", border=False):
                st.header(constants.NER_FEW_SHOT_PROMPTING_TAB_FORM_HEADER)
                system_message = st.text_area(
                    label=constants.TAB_FORM_SYSTEM_MESSAGE,
                    value=constants.NER_FEW_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    placeholder=constants.NER_FEW_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    height=200
                )
                ner_message = st.text_area(
                    label="NER (categories definition)",
                    value=constants.NER_FEW_SHOT_PROMPTING_TAB_CATEGORIES,
                    placeholder=constants.NER_FEW_SHOT_PROMPTING_TAB_CATEGORIES,
                    height=200
                )

                st.subheader("First example")
                col1, col2 = st.columns(2)

                with col1:
                    human_message_1 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        value=constants.NER_FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_1,
                        placeholder=constants.NER_FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_1,
                        height=200
                    )

                with col2:
                    ai_message_1 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        value=constants.NER_FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_1,
                        placeholder=constants.NER_FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_1,
                        height=200
                    )

                st.subheader("Actual email")
                human_message_2 = st.text_area(
                    label=constants.TAB_FORM_HUMAN_MESSAGE,
                    value=constants.NER_FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_2,
                    placeholder=constants.NER_FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_2,
                    height=200
                )
                submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON)

                if submitted:
                    if is_any_field_empty(
                            [system_message, ner_message, human_message_1, ai_message_1, human_message_2]):
                        st.warning(constants.TAB_FORM_EMPTY_FIELD_WARNING)
                        st.stop()

                    with get_openai_callback() as callbacks:
                        prompt = ChatPromptTemplate.from_messages([
                            SystemMessagePromptTemplate.from_template(system_message),
                            HumanMessage(human_message_1),
                            AIMessage(ai_message_1),
                            HumanMessage(human_message_2)
                        ])
                        chain = prompt | llm

                        with st.spinner("Generating response..."):
                            response = chain.invoke(input={"categories": ner_message})

                        st.markdown(constants.TAB_FORM_BOT_RESPONSE)
                        create_annotated_text(response.content)
                        st.markdown(constants.TAB_FORM_FULL_PROMPT)
                        st.code(prompt.format(categories=ner_message), language="text")
                        st.markdown(constants.TAB_FORM_REQUEST_STATS)
                        st.text(callbacks)
                        create_success_toast()


def create_ner_zero_shot_prompting_tab(ner_zero_shot_prompting: str, llm: AzureChatOpenAI) -> None:
    with ner_zero_shot_prompting:
        st.header(constants.NER_ZERO_SHOT_PROMPTING_TAB_HEADER)

        with st.expander(constants.TAB_DESCRIPTION_EXPANDER_TEXT, expanded=True):
            st.markdown(read_md_file("markdowns/ner-zero-shot-prompting-description.md"))

        with st.expander(constants.TAB_EXAMPLE_EXPANDER_TEXT):
            st.markdown(read_md_file("markdowns/ner-zero-shot-prompting-example.md"))

        with st.expander(constants.TAB_FORM_EXPANDER_TEXT):
            with st.form("ner_zero_shot_prompting_form", border=False):
                st.header(constants.NER_ZERO_SHOT_PROMPTING_TAB_FORM_HEADER)
                system_message = st.text_area(
                    label=constants.TAB_FORM_SYSTEM_MESSAGE,
                    value=constants.NER_ZERO_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    placeholder=constants.NER_ZERO_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    height=200
                )
                ner_message = st.text_area(
                    label="NER (categories definition)",
                    value=constants.NER_ZERO_SHOT_PROMPTING_TAB_CATEGORIES,
                    placeholder=constants.NER_ZERO_SHOT_PROMPTING_TAB_CATEGORIES,
                    height=200
                )
                human_message = st.text_area(
                    label=constants.TAB_FORM_HUMAN_MESSAGE,
                    value=constants.NER_ZERO_SHOT_PROMPTING_TAB_HUMAN_MESSAGE,
                    placeholder=constants.NER_ZERO_SHOT_PROMPTING_TAB_HUMAN_MESSAGE,
                    height=200
                )
                submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON)

                if submitted:
                    if is_any_field_empty([system_message, ner_message, human_message]):
                        st.warning(constants.TAB_FORM_EMPTY_FIELD_WARNING)
                        st.stop()

                    with get_openai_callback() as callbacks:
                        prompt = ChatPromptTemplate.from_messages([
                            SystemMessagePromptTemplate.from_template(system_message),
                            HumanMessage(human_message)
                        ])
                        chain = prompt | llm

                        with st.spinner("Generating response..."):
                            response = chain.invoke(input={"categories": ner_message})

                        st.markdown(constants.TAB_FORM_BOT_RESPONSE)
                        st.text(response.content)
                        st.markdown(constants.TAB_FORM_FULL_PROMPT)
                        st.code(prompt.format(categories=ner_message), language="text")
                        st.markdown(constants.TAB_FORM_REQUEST_STATS)
                        st.text(callbacks)
                        create_success_toast()


def create_few_shot_prompting_tab(few_shot_prompting_tab: str, llm: AzureChatOpenAI) -> None:
    with few_shot_prompting_tab:
        st.header(constants.FEW_SHOT_PROMPTING_TAB_HEADER)

        with st.expander(constants.TAB_DESCRIPTION_EXPANDER_TEXT, expanded=True):
            st.markdown(read_md_file("markdowns/few-shot-prompting-description.md"))

        with st.expander(constants.TAB_EXAMPLE_EXPANDER_TEXT):
            st.markdown(read_md_file("markdowns/few-shot-prompting-example.md"))

        with st.expander(constants.TAB_FORM_EXPANDER_TEXT):
            with st.form("few_shot_prompting_form", border=False):
                st.header(constants.FEW_SHOT_PROMPTING_TAB_FORM_HEADER)
                system_message = st.text_area(
                    label=constants.TAB_FORM_SYSTEM_MESSAGE,
                    value=constants.FEW_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    placeholder=constants.FEW_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    height=200
                )
                col1, col2, col3 = st.columns(3)

                with col1:
                    st.subheader("First example")
                    human_message_1 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        value=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_1,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_1,
                        height=200
                    )
                    ai_message_1 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        value=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_1,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_1,
                        height=25
                    )

                with col2:
                    st.subheader("Second example")
                    human_message_2 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        value=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_2,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_2,
                        height=200
                    )
                    ai_message_2 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        value=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_2,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_2,
                        height=25
                    )

                with col3:
                    st.subheader("Third example")
                    human_message_3 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        value=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_3,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_3,
                        height=200
                    )
                    ai_message_3 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        value=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_3,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_3,
                        height=25
                    )

                st.subheader("Below fields can be empty. They are optional.")
                col4, col5, col6 = st.columns(3)

                with col4:
                    st.subheader("Fourth example")
                    human_message_4 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_1,
                        height=200
                    )
                    ai_message_4 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_1,
                        height=25
                    )

                with col5:
                    st.subheader("Fifth example")
                    human_message_5 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_2,
                        height=200
                    )
                    ai_message_5 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_2,
                        height=25
                    )

                with col6:
                    st.subheader("Sixth example")
                    human_message_6 = st.text_area(
                        label=constants.TAB_FORM_HUMAN_MESSAGE,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_3,
                        height=200
                    )
                    ai_message_6 = st.text_area(
                        label=constants.TAB_FORM_AI_MESSAGE,
                        placeholder=constants.FEW_SHOT_PROMPTING_TAB_AI_MESSAGE_3,
                        height=25
                    )

                st.subheader("Actual email")
                human_message_7 = st.text_area(
                    label=constants.TAB_FORM_HUMAN_MESSAGE,
                    value=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_7,
                    placeholder=constants.FEW_SHOT_PROMPTING_TAB_HUMAN_MESSAGE_7,
                    height=200
                )
                submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON)

                if submitted:
                    if is_any_field_empty([system_message, human_message_1, ai_message_1, human_message_2, ai_message_2,
                                           human_message_3, ai_message_3, human_message_7]):
                        st.warning(constants.TAB_FORM_EMPTY_FIELD_WARNING)
                        st.stop()

                    with get_openai_callback() as callbacks:
                        prompt = ChatPromptTemplate.from_messages([
                            SystemMessage(system_message),
                            HumanMessage(human_message_1),
                            AIMessage(ai_message_1),
                            HumanMessage(human_message_2),
                            AIMessage(ai_message_2),
                            HumanMessage(human_message_3),
                            AIMessage(ai_message_3)
                        ])

                        if human_message_4 and ai_message_4:
                            prompt += HumanMessage(human_message_4) + AIMessage(ai_message_4)
                        if human_message_5 and ai_message_5:
                            prompt += HumanMessage(human_message_5) + AIMessage(ai_message_5)
                        if human_message_6 and ai_message_6:
                            prompt += HumanMessage(human_message_6) + AIMessage(ai_message_6)

                        prompt += HumanMessage(human_message_7)
                        chain = prompt | llm

                        with st.spinner("Generating response..."):
                            response = chain.invoke({})

                        st.markdown(constants.TAB_FORM_BOT_RESPONSE)

                        try:
                            st.json(json.loads(response.content))
                        except ValueError:
                            st.text(response.content)

                        st.markdown(constants.TAB_FORM_FULL_PROMPT)
                        st.code(prompt.format(), language="text")
                        st.markdown(constants.TAB_FORM_REQUEST_STATS)
                        st.text(callbacks)
                        create_success_toast()

        with st.expander(constants.TAB_STATS_EXPANDER_TEXT):
            st.markdown(read_md_file("markdowns/stats-description.md"))
            create_stats("few-shot-prompting")


def create_zero_shot_prompting_tab(zero_shot_prompting_tab: str, llm: AzureChatOpenAI) -> None:
    with zero_shot_prompting_tab:
        st.header(constants.ZERO_SHOT_PROMPTING_TAB_HEADER)

        with st.expander(constants.TAB_DESCRIPTION_EXPANDER_TEXT, expanded=True):
            st.markdown(read_md_file("markdowns/zero-shot-prompting-description.md"))

        with st.expander(constants.TAB_EXAMPLE_EXPANDER_TEXT):
            st.markdown(read_md_file("markdowns/zero-shot-prompting-example.md"))

        with st.expander(constants.TAB_FORM_EXPANDER_TEXT):
            with st.form("zero_shot_prompting_form", border=False):
                st.header(constants.ZERO_SHOT_PROMPTING_TAB_FORM_HEADER)
                system_message = st.text_area(
                    label=constants.TAB_FORM_SYSTEM_MESSAGE,
                    value=constants.ZERO_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    placeholder=constants.ZERO_SHOT_PROMPTING_TAB_SYSTEM_MESSAGE,
                    height=200
                )
                human_message = st.text_area(
                    label=constants.TAB_FORM_HUMAN_MESSAGE,
                    value=constants.ZERO_SHOT_PROMPTING_TAB_HUMAN_MESSAGE,
                    placeholder=constants.ZERO_SHOT_PROMPTING_TAB_HUMAN_MESSAGE,
                    height=200
                )
                submitted = st.form_submit_button(label=constants.TAB_FORM_SUBMIT_BUTTON)

                if submitted:
                    if is_any_field_empty([system_message, human_message]):
                        st.warning(constants.TAB_FORM_EMPTY_FIELD_WARNING)
                        st.stop()

                    with get_openai_callback() as callbacks:
                        prompt = ChatPromptTemplate.from_messages([
                            SystemMessage(system_message),
                            HumanMessage(human_message)
                        ])
                        chain = prompt | llm

                        with st.spinner("Generating response..."):
                            response = chain.invoke({})

                        st.markdown(constants.TAB_FORM_BOT_RESPONSE)

                        try:
                            st.json(json.loads(response.content))
                        except ValueError:
                            st.text(response.content)

                        st.markdown(constants.TAB_FORM_FULL_PROMPT)
                        st.code(prompt.format(), language="text")
                        st.markdown(constants.TAB_FORM_REQUEST_STATS)
                        st.text(callbacks)
                        create_success_toast()


def create_stats(approach: str):
    data = pd.read_excel("important-files/confusion-matrix.xlsx", sheet_name=f"{approach}-cm")
    result = {}
    for index, row in data.iterrows():
        result.update({
            row['field_name']: {
                'origin_location': row['origin_location'],
                'destination_location': row['destination_location'],
                'currency': row['currency'],
                'distance': row['distance'],
                'transport_type': row['transport_type'],
                'trailer_type': row['trailer_type'],
                'dangerous_hazardous': row['dangerous_hazardous'],
                'cargo_type': row['cargo_type'],
                'steps': row['steps'],
                'validity_date': row['validity_date'],
                'loading_date': row['loading_date'],
                'arrival_date': row['arrival_date'],
                'weight': row['weight'],
                'volume': row['volume'],
                'pallets': row['pallets'],
                'vehicle_loading_method': row['vehicle_loading_method'],
                'temperature_requirements': row['temperature_requirements'],
                'seals': row['seals'],
                'others': row['others']
            }
        })
    cm = ConfusionMatrix(matrix=result)

    if not os.path.exists(f"stats/{approach}-stat.pycm"):
        cm.save_stat(
            f"stats/{approach}-stat",
            overall_param=["ACC Macro", "F1 Macro", "Kappa", "NPV Macro", "Overall ACC", "PPV Macro",
                           "SOA1(Landis & Koch)", "TPR Macro", "zero-one Loss"],
            class_param=["ACC", "AUC", "AUCI", "F1", "FN", "FP", "FPR", "N", "P", "POP", "PPV", "TN", "TON", "TOP",
                         "TP", "TPR"]
        )

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.subheader("Class statistics")
        st.dataframe(
            pd.read_excel(f"stats/{approach}-stat-class.xlsx").astype(str), hide_index=True,
            use_container_width=True
        )

    with col2:
        st.subheader("Overall statistics")
        st.dataframe(
            pd.read_excel(f"stats/{approach}-stat-overall.xlsx").astype(str), hide_index=True,
            use_container_width=True
        )

    cm.plot(cmap=plt.cm.Reds, number_label=True, plot_lib="matplotlib")
    plt.xticks(rotation=90)
    st.pyplot(plt, use_container_width=False)


def get_cosine_similarity(vec1: Counter, vec2: Counter) -> float:
    intersection = set(vec1.keys()) & set(vec2.keys())
    numerator = sum([vec1[x] * vec2[x] for x in intersection])

    sum1 = sum([vec1[x] ** 2 for x in list(vec1.keys())])
    sum2 = sum([vec2[x] ** 2 for x in list(vec2.keys())])
    denominator = math.sqrt(sum1) * math.sqrt(sum2)

    if not denominator:
        return 0.0
    else:
        return float(numerator) / denominator


def text_to_vector(text: str) -> Counter:
    words = re.compile(r"\w+").findall(text)
    return Counter(words)


def create_sheet_for_cm(saved_excel_path: str, actual_sheet_name: str, predicted_sheet_name: str,
                        timestamp: str) -> None:
    actual_dataframe = pd.read_excel(saved_excel_path, actual_sheet_name)
    predicted_dataframe = pd.read_excel(saved_excel_path, predicted_sheet_name)
    diff_dataframe = predicted_dataframe[predicted_dataframe != actual_dataframe]
    differences = {}

    for row_index in range(len(diff_dataframe)):
        diff_col_and_value = diff_dataframe.iloc[row_index].dropna()
        col_name = diff_col_and_value.index

        for value in range(len(diff_dataframe.iloc[row_index][col_name].values)):
            diff_text = str(diff_dataframe.iloc[row_index][col_name].values[value])
            actual_text = str(actual_dataframe.iloc[row_index][col_name].values[value])
            diff_vector = text_to_vector(diff_text)
            actual_vector = text_to_vector(actual_text)
            cosine = get_cosine_similarity(diff_vector, actual_vector)

            if cosine >= 0.8:
                differences.update({row_index: diff_text})

    for row_index in differences:
        for value in predicted_dataframe.iloc[row_index].values:
            if value == differences[row_index]:
                column_index = list(predicted_dataframe.iloc[row_index].values).index(value)
                actual_value = actual_dataframe.iloc[row_index].values[column_index]
                predicted_dataframe.iloc[row_index, column_index] = actual_value

    with pd.ExcelWriter(
            saved_excel_path,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="replace",
    ) as writer:
        predicted_dataframe.to_excel(writer, sheet_name=timestamp, index=False)
        writer.book.save(filename=saved_excel_path)


def create_cm(excel_file_path: str, actual_sheet_name: str, timestamp: str) -> None:
    actual_dataframe = pd.read_excel(
        excel_file_path,
        sheet_name=actual_sheet_name
    )
    predicted_dataframe = pd.read_excel(
        excel_file_path,
        sheet_name=timestamp
    )

    actual_dataframe = actual_dataframe.drop(["id_from_dataset", "llm_input"], axis=1)
    predicted_dataframe = predicted_dataframe.drop(["id_from_dataset", "llm_input"], axis=1)

    result_matrix = np.zeros((len(actual_dataframe.columns), len(actual_dataframe.columns)))
    diagonal = []

    for col in actual_dataframe.columns:
        match_amount = actual_dataframe[col].isin(predicted_dataframe[col]).value_counts()[True]
        diagonal.append(match_amount)

    np.fill_diagonal(result_matrix, diagonal)
    last_column = [row[-1] for row in result_matrix]

    for (index, val) in enumerate(last_column):
        last_column[index] = len(actual_dataframe) - diagonal[index]
        result_matrix.itemset((index, -1), last_column[index])

    result = {}
    for index, row in actual_dataframe.iterrows():
        if index == 19:
            break
        result.update({
            actual_dataframe.columns.values[index]: {
                'origin_location': int(result_matrix[index][0]),
                'destination_location': int(result_matrix[index][1]),
                'currency': int(result_matrix[index][2]),
                'distance': int(result_matrix[index][3]),
                'transport_type': int(result_matrix[index][4]),
                'trailer_type': int(result_matrix[index][5]),
                'dangerous_hazardous': int(result_matrix[index][6]),
                'cargo_type': int(result_matrix[index][7]),
                'steps': int(result_matrix[index][8]),
                'validity_date': int(result_matrix[index][9]),
                'loading_date': int(result_matrix[index][10]),
                'arrival_date': int(result_matrix[index][11]),
                'weight': int(result_matrix[index][12]),
                'volume': int(result_matrix[index][13]),
                'pallets': int(result_matrix[index][14]),
                'vehicle_loading_method': int(result_matrix[index][15]),
                'temperature_requirements': int(result_matrix[index][16]),
                'seals': int(result_matrix[index][17]),
                'others': int(result_matrix[index][18])
            }
        })

    cm = ConfusionMatrix(matrix=result)
    cm.plot(cmap=plt.cm.Reds, number_label=True, plot_lib="matplotlib")
    plt.xticks(rotation=90)
    st.pyplot(plt, use_container_width=False)

    remove_excel_file(excel_file_path)


def remove_excel_sheet(filename: str, sheet_name: str) -> None:
    try:
        work_book = openpyxl.load_workbook(filename)
        sheet_to_remove = work_book[sheet_name]
        work_book.remove(sheet_to_remove)
        work_book.save(filename)
    except:
        st.error("Sheet does not exist!")


def remove_excel_file(excel_file_path: str) -> None:
    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)


def save_excel_file(excel_file) -> str:
    to_save = openpyxl.load_workbook(io.BytesIO(excel_file.read()))
    to_save.save(f"uploaded-files/{excel_file.name}")
    to_save.close()
    return f"uploaded-files/{excel_file.name}"


def create_success_toast() -> None:
    st.toast("Success!", icon="🎉")


def create_annotated_text(text: str) -> None:
    entities = re.findall(r"\[(.*?)\]\((.*?)\)", text)
    words = re.split("\[(.*?)\)", text)

    for word in words:
        for entity in entities:
            if entity[0] in word and entity[1] in word:
                words[words.index(word)] = entity

    annotated_text(words)


def is_any_field_empty(list_of_fields: list) -> bool:
    return any([field == "" for field in list_of_fields])


def create_azure_openai_model(openai_api_key: str, temperature: float) -> AzureChatOpenAI:
    return AzureChatOpenAI(
        azure_endpoint="https://open-ai-resource-gen-ai.openai.azure.com/",
        openai_api_version="2023-07-01-preview",
        openai_api_key=openai_api_key,
        openai_api_type="azure",
        deployment_name="gpt-35-dev",
        model_name="gpt-35-turbo",
        model_version="0613",
        temperature=temperature
    )


def create_azure_openai_embedding_model(openai_api_key: str) -> AzureOpenAIEmbeddings:
    return AzureOpenAIEmbeddings(
        azure_endpoint="https://open-ai-resource-gen-ai.openai.azure.com/",
        openai_api_version="2023-07-01-preview",
        openai_api_key=openai_api_key,
        openai_api_type="azure"
    )


def create_page_config() -> None:
    st.set_page_config(
        page_title="Email to Quote 📧➡️💰",
        page_icon="📧",
        layout="wide",
        initial_sidebar_state="collapsed",
        menu_items={
            "Get Help": "https://www.transporeon.com/en",
            "Report a bug": "https://www.transporeon.com/en",
            "About": constants.MAIN_MENU_ABOUT
        }
    )


def hide_sidebar() -> None:
    st.markdown(constants.HIDE_SIDEBAR_AND_DEPLOY_HTML, unsafe_allow_html=True)


def display_sidebar() -> None:
    st.markdown(constants.DISPLAY_SIDEBAR_HTML, unsafe_allow_html=True)


def create_authenticator() -> stauth.Authenticate:
    with open("authentication.yaml") as file:
        config = yaml.load(file, Loader=SafeLoader)

    return stauth.Authenticate(
        config["credentials"],
        config["cookie"]["name"],
        config["cookie"]["key"],
        config["cookie"]["expiry_days"],
        config["preauthorized"]
    )


def read_md_file(file_path: str) -> str:
    return open(file_path, encoding="utf8").read()


if __name__ == "__main__":
    main()
